import datetime
from kivy.app import App
from kivy.lang import Builder
from kivy.uix.screenmanager import ScreenManager, Screen
from openpyxl import load_workbook

price_dic = {
    "coffee": 2500,
    "tea": 3000,
    "3 in 1": 5000,
    "ice 3 in 1": 7000,
    "ice coffee": 6000,
    "ice tea": 6000,
    "4 cheese": 20000,
    "hot dog": 10000
}

wb = load_workbook("orders.xlsx")
sheet_1 = wb['active_tables']
sheet_2 = wb['todays_orders']

active_tables_nums = []


class AddOrders(Screen):
    order_type_list = ["Hot", "Cold", "Food"]
    order_choice = []
    order_list = []
    var_1 = None

    def table_num_input(self):
        table_num = self.ids.tab_num.text
        table_num = table_num.zfill(2)
        now = datetime.datetime.now()
        timestamp = now.strftime("%d/%m/%Y %H:%M:%S")
        self.ids.submit_order_input.disabled = True
        if table_num != "00":
            self.ids.show_table_num.text = f'Table Number: {table_num}'
        if table_num != "00":
            self.order_list.append(table_num)
            self.order_list.append(timestamp)
            self.ids.tab_num.text = ""
            self.ids.tab_num.disabled = True

    def order_type(self, value):
        global order_choice
        match value:
            case "Hot":
                order_choice = ["3 in 1", "coffee", "tea"]
                self.ids.order_id.values = order_choice

            case "Cold":
                order_choice = ["ice 3 in 1", "ice coffee", "ice tea"]
                self.ids.order_id.values = order_choice

            case "Food":
                order_choice = ["4 cheese", "hot dog"]
                self.ids.order_id.values = order_choice

    def order_choice(self, value):
        global var_1
        var_1 = value

    def order_save(self):
        global var_1
        self.order_list.append(var_1)
        var_1 = None
        self.ids.order_type_id.text = "Select Order Type"
        self.ids.order_id.text = "Select Order"
        self.ids.submit_order_input.disabled = False

    def submit_orders(self):
        self.order_list = [x for x in self.order_list if x != "Select Order"]

        num_rows = sheet_1.max_row
        for i, item in enumerate(self.order_list, start=1):
            sheet_1.cell(row=num_rows + 1, column=i).value = item
        wb.save("orders.xlsx")

    def clear_everything(self):
        self.order_list = []
        self.ids.tab_num.text = ""
        self.ids.order_type_id.text = "Select Order Type"
        self.ids.order_id.text = "Select Order"
        self.ids.tab_num.disabled = False
        self.ids.show_table_num.text = ""
        self.ids.show_orders.text = ""

    def show_orders(self):
        show_orders_str = " "
        for i in self.order_list[2:]:
            show_orders_str = str(show_orders_str) + str(i) + " || "
            self.ids.show_orders.text = f'orders: {show_orders_str}'

    def cancel_order_list(self):
        self.ids.cancel_order_list.values = self.order_list[2:]

    def cancel_order(self):
        order_to_cancel = self.ids.cancel_order_list.text
        self.order_list.remove(order_to_cancel)
        self.ids.cancel_order_list.text = "Choose Order to Cancel"

    def find_checkout_table(self):
        global active_tables_nums
        for row in sheet_1.iter_rows(min_col=1, max_col=1, min_row=2, values_only=True):
            for cell in row:
                active_tables_nums.append(cell)


class CheckoutPage(Screen):
    index_nums = []
    checkout_order_list = []
    checkout_price_list = []
    grand_total = 0
    checkout_append_list = []
    receipt_orders = []
    receipt_iterations = []
    receipt_orders_mod = []
    unit_total_list = []
    active_tables_nums_checkout = []

    def create_active_tables_nums_list(self):
        for row in sheet_1.iter_rows(min_col=1, max_col=1, min_row=2, values_only=True):
            for cell in row:
                self.active_tables_nums_checkout.append(cell)

    def spinner_checkout_list_show(self):
        self.ids.spinner_checkout_list.values = sorted(set(self.active_tables_nums_checkout))
        self.checkout_order_list = []
        self.checkout_price_list = []
        self.index_nums = []
        self.grand_total = 0
        self.checkout_append_list = []
        self.ids.checkout_submit_button.disabled = False

    def find_index_nums(self):
        tab_num = self.ids.spinner_checkout_list.text
        n = 2
        now = datetime.datetime.now()
        timestamp2 = now.strftime("%d/%m/%Y %H:%M:%S")
        self.checkout_append_list.append(tab_num)
        self.checkout_append_list.append(timestamp2)
        for num in active_tables_nums:
            if num == tab_num:
                self.index_nums.append(n)
                n += 1
            else:
                n += 1

    def create_order_list_for_checkout(self):
        for number in self.index_nums:
            for column in sheet_1.iter_rows(min_row=number, max_row=number, min_col=3, values_only=True):
                for order in column:
                    self.checkout_order_list.append(order)
                    self.checkout_order_list = [x for x in self.checkout_order_list if x is not None]

    def show_labels(self):
        self.ids.show_the_word_orders.text = "The Orders Are:"
        show_the_orders_string = ""
        for item in self.checkout_order_list:
            show_the_orders_string = str(show_the_orders_string) + str(item) + " || "
            self.ids.show_order_list_string.text = show_the_orders_string

        self.checkout_append_list.append(self.grand_total)

    def create_eq_price_list(self):
        for i in self.checkout_order_list:
            eq = price_dic.get(i)
            self.checkout_price_list.append(eq)
        for price in self.checkout_price_list:
            self.grand_total = self.grand_total + int(price)
        self.ids.show_grand_total.text = f'Grand Total: {str(self.grand_total)}'
        self.ids.checkout_submit_button.disabled = True

    # here starts after the checkout_checkout_checkout button is pressed:
    # this function deletes the row from the xlsx:
    def delete_checkout_row_from_active(self):
        for dr in sorted(self.index_nums, reverse=True):
            sheet_1.delete_rows(dr)
        wb.save("orders.xlsx")

    def input_checkout_append_list_to_sheet2(self):
        num_row = sheet_2.max_row
        for i, item in enumerate(self.checkout_append_list, start=1):
            sheet_2.cell(row=num_row + 1, column=i).value = item
        wb.save("orders.xlsx")

    def create_receipt_orders(self):
        for qwe in self.checkout_order_list:
            if qwe not in self.receipt_orders:
                self.receipt_orders.append(qwe)
                iteration = self.checkout_order_list.count(qwe)
                self.receipt_iterations.append(iteration)

    def create_receipt_list_mod(self):
        for qwe in self.receipt_orders:
            while len(qwe) < 13:
                qwe = qwe + " "
            self.receipt_orders_mod.append(qwe)

    def create_unit_total_list(self):
        for i, unit in enumerate(self.receipt_orders):
            unit_total = price_dic.get(unit) * self.receipt_iterations[i]
            self.unit_total_list.append(unit_total)

    def create_receipt(self):
        table_number_final = self.ids.spinner_checkout_list.text
        now = datetime.datetime.now()
        time = now.strftime("%d/%m/%Y %H:%M:%S")
        header = "Date: " + time + "          " + "Table Number: " + str(table_number_final) + "\n" + "\n"
        next_line = "Orders: " + "          " + "Amount: " + "          " + "Unit Total: " + "\n" + "\n"
        septum = "----------------------------------------------------------------------------" + "\n" + "\n"
        with open("receipt.txt", "a") as receipt_file:
            receipt_file.write(header)
            receipt_file.write(next_line)
            for i, unit in enumerate(self.receipt_orders_mod):
                unit_total_receipt = unit + "     " + str(self.receipt_iterations[i]) + "                 " + str(
                    self.unit_total_list[i]) + "\n"
                receipt_file.write(unit_total_receipt)
            grand_total = 0
            for price in self.unit_total_list:
                grand_total = int(grand_total) + int(price)
            grand_total_receipt = "\n" + "              Grand Total = " + str(grand_total) + "\n"
            receipt_file.write(grand_total_receipt)
            receipt_file.write(septum)

    def clear_everything(self):
        global active_tables_nums
        self.active_tables_nums_checkout = []
        self.index_nums = []
        self.checkout_order_list = []
        self.checkout_price_list = []
        self.grand_total = 0
        self.checkout_append_list = []
        self.receipt_orders = []
        self.receipt_iterations = []
        self.receipt_orders_mod = []
        self.unit_total_list = []
        self.ids.show_the_word_orders.text = ""
        self.ids.show_order_list_string.text = ""
        self.ids.show_grand_total.text = ""
        self.ids.spinner_checkout_list.text = "Choose Table for Checkout"


class ActiveTabsPage(Screen):
    active_tables_nums_check_active = []
    index_nums = []
    col_list = []
    order_to_delete_list = []
    m = 1

    def create_active_tables_nums_list_check_active(self):
        for row in sheet_1.iter_rows(min_col=1, max_col=1, min_row=2, values_only=True):
            for cell in row:
                self.active_tables_nums_check_active.append(cell)

    def spinner_fill(self):
        self.ids.spinner_check_active_list.values = sorted(set(self.active_tables_nums_check_active))

    def create_index_nums_list(self):
        tab_num = self.ids.spinner_check_active_list.text
        self.ids.order_rows.text = f'Click Here To View The Log Of Table: {tab_num}'
        self.ids.submit_deleting_order.disabled = False
        self.ids.order_to_delete.disabled = False
        n = 2
        for num in active_tables_nums:
            if num == tab_num:
                self.index_nums.append(n)
                n += 1
            else:
                n += 1

    def print_the_order_rows(self):
        for i in self.index_nums:
            for col in sheet_1.iter_rows(min_row=i, max_row=i, min_col=2, values_only=True):
                if col is not None:
                    col = str(col)
                    col = col.replace("(", "")
                    col = col.replace(")", "")
                    col = col.replace(", None", "")
                    self.col_list.append(col)
            self.ids.order_rows.values = self.col_list

    def clear_everything(self):
        self.active_tables_nums_check_active = []
        self.col_list = []
        self.index_nums = []
        self.order_to_delete_list = []
        self.ids.order_to_delete.text = "Choose an Order To Delete"
        self.ids.order_to_delete.disabled = True
        self.ids.order_rows.text = ""
        self.ids.order_rows.values = []
        self.ids.submit_deleting_order.disabled = True
        self.ids.spinner_check_active_list.text = "Active Tables"

    def spinner_delete_order(self):
        for i in self.index_nums:
            for col in sheet_1.iter_rows(min_row=i, max_row=i, min_col=3, values_only=True):
                for item in col:
                    if item is not None:
                        item = str(item)
                        self.order_to_delete_list.append(item)
        self.ids.order_to_delete.values = set(self.order_to_delete_list)

    def delete_order(self):
        order_to_delete = self.ids.order_to_delete.text
        print(self.index_nums, order_to_delete)
        for i in self.index_nums:
            print(i)
            for orders in sheet_1.iter_rows(min_row=i, max_row=i, min_col=3, values_only=True):
                print(orders)
                for x, item in enumerate(orders):
                    if item == order_to_delete:
                        x = x + 3
                        sheet_1.cell(row=i, column=x).value = ""
                    wb.save("orders.xlsx")
                    break


class LogOutPage(Screen):
    active_tables_nums_logout = []
    grand_grand_total = 0

    def exit_function(self):
        if self.ids.confirmation_text.text == "obl":
            for row in sheet_1.iter_rows(min_col=1, max_col=1, min_row=2, values_only=True):
                for cell in row:
                    self.active_tables_nums_logout.append(cell)
            if len(self.active_tables_nums_logout) == 0:
                self.calculate_grand_grand_total()
                self.ids.exit_button.disabled = False
                self.ids.error_clear_tables.text = "Your Grand Total For Today Is"
                self.ids.grand_grand_total_text.text = str(self.grand_grand_total)
            else:
                self.ids.error_clear_tables.text = "You Must Checkout All Tables Before You Logout"

    def calculate_grand_grand_total(self):
        for total in sheet_2.iter_rows(min_col=3, max_col=3, min_row=2, values_only=True):
            total = str(total)
            total = total.replace("(", "")
            total = total.replace(")", "")
            total = total.replace("'", "")
            total = total.replace(",", "")
            total = total.replace('"', '')
            self.grand_grand_total = int(self.grand_grand_total) + int(total)

    def clear_everything(self):
        self.active_tables_nums_logout = []
        self.ids.exit_button.disabled = True
        self.ids.confirmation_text.text = ""
        self.ids.error_clear_tables.text = ""
        self.ids.grand_grand_total_text.text = ""
        self.grand_grand_total = 0

    def clear_sheet_2(self):
        max = sheet_2.max_row
        print("1", max)
        row_nums_clear = [1, 2, 3]
        for dr in reversed(row_nums_clear):
            sheet_2.delete_cols(dr)
            print("2", max)
            wb.save("orders.xlsx")
        print("3", max)


class WindowManager(ScreenManager):
    pass


kv = Builder.load_file("cafeorders.kv")


class MyApp(App):
    def build(self):
        return kv


if __name__ == "__main__":
    MyApp().run()
